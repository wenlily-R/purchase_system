# -*- coding: utf-8 -*-
"""V11.1 三方询价升级验证: 至少3家校验 + Excel导出(选中高亮/最低价/决策备注)"""
import os, sys, sqlite3, json, shutil, time, io
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import app as S

# 屏蔽真实钉钉
S.start_instances = lambda *a, **k: None

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'purchase.db')
conn = sqlite3.connect(DB, timeout=30)
MARK = '【询价导出验证】'
# 清理历史测试
for t in ('inquiry_suppliers', 'inquiries'):
    pass
# 构造已通过申请
conn.execute("INSERT INTO purchase_requests(req_no,status,purpose,dept,requester,requester_id) VALUES('XQTEST'||strftime('%s'),'已通过',?,'综合办','测试员',1)", (MARK,))
req_id = conn.execute("SELECT id FROM purchase_requests WHERE purpose=?", (MARK,)).fetchone()[0]
conn.execute("INSERT INTO request_items(req_id,item_name,spec,unit,quantity,total_price) VALUES(?,?,?,?,?,?)", (req_id, '测试物料A', '规格X', '个', 2, 2000))
conn.execute("INSERT INTO request_items(req_id,item_name,spec,unit,quantity,total_price) VALUES(?,?,?,?,?,?)", (req_id, '测试物料B', '规格Y', '箱', 3, 3000))
conn.commit()
print('构造申请#%s 明细2项' % req_id)

client = S.app.test_client()
r = client.post('/api/login', json={'username': 'admin', 'password': 'admin123'})
assert r.status_code == 200, '登录失败'

# ① 不足3家应被拦截
r = client.post('/api/inquiries', json={'req_id': req_id, 'suppliers': [{'name': '甲供应商'}, {'name': '乙供应商'}]})
assert r.status_code == 400, '不足3家未被拦截!'
assert '3家及以上' in r.get_json()['error'], r.get_json()
print('① 不足3家拦截 OK:', r.get_json()['error'])

# ② 满3家可创建
sups = [{'name': '甲供应商', 'contact': '张三', 'phone': '13800000001'},
        {'name': '乙供应商', 'contact': '李四', 'phone': '13800000002'},
        {'name': '丙供应商', 'contact': '王五', 'phone': '13800000003'}]
r = client.post('/api/inquiries', json={'req_id': req_id, 'suppliers': sups})
assert r.status_code == 200, r.get_data(as_text=True)
iid = r.get_json()['id']
inq_no = r.get_json()['inq_no']
print('② 创建询价单 OK:', inq_no, 'id=', iid)

# ③ 供应商报价(甲乙报价, 丙未报价)
toks = [row[0] for row in conn.execute("SELECT token FROM inquiry_suppliers WHERE inquiry_id=?", (iid,)).fetchall()]
assert len(toks) == 3
r = client.post('/api/inquiry/vendor/%s/quote' % toks[0], json={'quote_price': 4500, 'quote_remark': '交期7天'})
assert r.status_code == 200
r = client.post('/api/inquiry/vendor/%s/quote' % toks[1], json={'quote_price': 4200, 'quote_remark': '交期5天'})
assert r.status_code == 200
print('③ 甲乙报价 OK (4500/4200), 丙未报价')

# ④ 选中乙(最低价4200)
sup2 = conn.execute("SELECT id FROM inquiry_suppliers WHERE inquiry_id=? AND supplier_name='乙供应商'", (iid,)).fetchone()[0]
r = client.post('/api/inquiries/%d/select' % iid, json={'supplier_id': sup2})
assert r.status_code == 200, r.get_data(as_text=True)
print('④ 选中乙供应商 OK, 生成订单:', r.get_json().get('order_no'))

# ⑤ 导出Excel
r = client.get('/api/inquiries/%d/export' % iid)
assert r.status_code == 200, r.get_data(as_text=True)[:200]
data = r.data
assert data[:2] == b'PK', '不是xlsx文件'
print('⑤ 导出Excel OK, 大小:', len(data), 'bytes')

# 检查Excel内容
from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(data))
ws = wb['询价单']
all_rows = [[c.value for c in row] for row in ws.iter_rows()]
flat = [str(v) for row in all_rows for v in row if v is not None]
joined = '|'.join(flat)
checks = {
    '询价单号': inq_no in joined,
    '标题': '采 购 询 价 单' in joined,
    '三家供应商': all(n in joined for n in ('甲供应商', '乙供应商', '丙供应商')),
    '最低价标记': '★ 最低价' in joined,
    '选中标记': '✅ 已选中' in joined,
    '决策备注': '多方询价' in joined and '合作方' in joined,
    '备注含交期': '交期5天' in joined,
}
print('⑥ Excel内容检查:')
for k, v in checks.items():
    print('   ', '✅' if v else '❌', k)

# 高亮检查: 乙供应商行应浅黄填充+标红
sel_row = None
for ri, row in enumerate(all_rows, 1):
    if row and row[1] == '乙供应商':
        sel_row = ri
        break
assert sel_row, '未找到乙供应商行'
fill_ok = ws.cell(sel_row, 2).fill.fgColor.rgb in ('00FFF2CC', 'FFF2CC')
font_ok = ws.cell(sel_row, 2).font.bold and ws.cell(sel_row, 2).font.color.rgb in ('00C00000', 'FFC00000')
print('    ', '✅' if fill_ok else '❌', '选中行浅黄高亮', ws.cell(sel_row, 2).fill.fgColor.rgb)
print('    ', '✅' if font_ok else '❌', '选中行加粗标红', ws.cell(sel_row, 2).font.bold, ws.cell(sel_row, 2).font.color.rgb)

# 最低价行(乙)标红加粗
min_ok = ws.cell(sel_row, 4).font.bold and ws.cell(sel_row, 4).font.color.rgb in ('00C00000', 'FFC00000')
print('    ', '✅' if min_ok else '❌', '报价金额标红加粗')

all_pass = all(checks.values()) and fill_ok and font_ok and min_ok
print('=== 结果:', 'ALL PASS' if all_pass else 'FAIL', '===')

# 清理测试数据
for t in ('inquiry_suppliers', 'inquiries'):
    pass
conn.execute("DELETE FROM inquiry_suppliers WHERE inquiry_id=?", (iid,))
conn.execute("DELETE FROM inquiries WHERE id=?", (iid,))
conn.execute("DELETE FROM order_items WHERE order_id IN (SELECT id FROM purchase_orders WHERE req_id=?)", (req_id,))
conn.execute("DELETE FROM approval_instances WHERE biz_type='purchase_order' AND biz_id IN (SELECT id FROM purchase_orders WHERE req_id=?)", (req_id,))
conn.execute("DELETE FROM purchase_orders WHERE req_id=?", (req_id,))
conn.execute("DELETE FROM request_items WHERE req_id=?", (req_id,))
conn.execute("DELETE FROM purchase_requests WHERE id=?", (req_id,))
conn.execute("DELETE FROM logs WHERE detail LIKE ?", ('%' + inq_no + '%',))
conn.execute("UPDATE sqlite_sequence SET seq=COALESCE((SELECT MAX(id) FROM inquiries),0) WHERE name='inquiries'")
conn.execute("UPDATE sqlite_sequence SET seq=COALESCE((SELECT MAX(id) FROM purchase_orders),0) WHERE name='purchase_orders'")
conn.commit()
conn.close()
print('清理完成')
sys.exit(0 if all_pass else 1)
