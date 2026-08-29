#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""V11.126 端到端验证 v2:
 询价1: 商家行报价(交付/质保汇总) → 最后一家报价触发自动定标审批 → 审批通过 → 订单生效
 询价2: SQL造报价 → 手动提交审批 → 驳回恢复询价中 → 重新提交(草稿去重)
 订单进度(红黄绿) / 导出Excel交付质保列 / 轮询存钉钉表单值
测试数据统一带【V11.126验证】标记, monkeypatch start_instances 屏蔽真实钉钉/飞书推送"""
import json, sqlite3, sys, os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import app as appmod

MARK = '【V11.126验证】'
PASS, FAIL = [], []

def check(name, cond, extra=''):
    (PASS if cond else FAIL).append(name)
    print(('  ✅ ' if cond else '  ❌ ') + name + ((' | ' + str(extra)) if extra else ''))

CALLED = {'start_instances': 0}
def _fake_start(*a, **k):
    CALLED['start_instances'] += 1
    return None
appmod.start_instances = _fake_start

def q(sql, args=()):
    c = sqlite3.connect('data/purchase.db', timeout=30)
    c.row_factory = sqlite3.Row
    try:
        return c.execute(sql, args).fetchall()
    finally:
        c.close()

def q1(sql, args=()):
    r = q(sql, args)
    return r[0] if r else None

client = appmod.app.test_client()
client.post('/api/login', json={'username': 'admin', 'password': 'admin123'})

# ══ 准备: 2 张测试申请(已通过) ══
c = sqlite3.connect('data/purchase.db', timeout=30)
cur = c.cursor()
cur.execute("INSERT INTO purchase_requests(req_no, dept, purpose, requester, requester_id, status, total_estimated, created_at, req_type) VALUES('V126T1', '后勤部', ?, '测试申请人', 1, '已通过', 100, datetime('now','localtime'), '物资采购')", (MARK + '申请1',))
rid1 = cur.lastrowid
cur.execute("INSERT INTO purchase_requests(req_no, dept, purpose, requester, requester_id, status, total_estimated, created_at, req_type) VALUES('V126T2', '后勤部', ?, '测试申请人', 1, '已通过', 100, datetime('now','localtime'), '物资采购')", (MARK + '申请2',))
rid2 = cur.lastrowid
cur.execute("INSERT INTO request_items(req_id, item_name, spec, unit, quantity, estimated_price, total_price) VALUES(?,?,?,?,?,?,?)", (rid1, '测试轴承', '6205', '个', 5, 20, 100))
cur.execute("INSERT INTO request_items(req_id, item_name, spec, unit, quantity, estimated_price, total_price) VALUES(?,?,?,?,?,?,?)", (rid1, '测试螺丝', 'M8', '个', 10, 5, 50))
cur.execute("INSERT INTO request_items(req_id, item_name, spec, unit, quantity, estimated_price, total_price) VALUES(?,?,?,?,?,?,?)", (rid2, '测试垫片', '10mm', '个', 4, 10, 40))
c.commit(); c.close()

# ══════════════ 询价1: 行报价 → 自动定标 → 通过 ══════════════
r = client.post('/api/inquiries', json={'req_id': rid1, 'suppliers': [
    {'name': 'V126厂家甲', 'contact': '甲', 'phone': '10000000001'},
    {'name': 'V126厂家乙', 'contact': '乙', 'phone': '10000000002'}]})
check('创建询价单1', r.status_code == 200 and r.get_json().get('success'), r.get_json())
inq1 = r.get_json()['inq_no']
iid1 = q1("SELECT id FROM inquiries WHERE inq_no=?", (inq1,))['id']
sups1 = q("SELECT * FROM inquiry_suppliers WHERE inquiry_id=?", (iid1,))
tokA = sups1[0]['token']; sidA = sups1[0]['id']; sidB = sups1[1]['id']

details = [
    {'unit_price': 30, 'qty': 5, 'delivery': '7天', 'warranty': '3个月', 'brand': '哈轴', 'remark': '含运'},
    {'unit_price': 8, 'qty': 10, 'delivery': '7天', 'warranty': '3个月', 'brand': '哈轴', 'remark': '含税'},
]
r = client.post('/api/inquiry/vendor/%s/quote' % tokA, json={'quote_price': 0, 'details': details})
check('商家甲行报价提交', r.status_code == 200 and r.get_json().get('success'), r.get_json())
sA = q1("SELECT * FROM inquiry_suppliers WHERE id=?", (sidA,))
check('报价合计=Σ单价×数量(230)', abs(float(sA['quote_price']) - 230.0) < 0.01, sA['quote_price'])
check('交付日期汇总存库(7天)', sA['quote_delivery'] == '7天', sA['quote_delivery'])
check('质保时间汇总存库(3个月)', sA['quote_warranty'] == '3个月', sA['quote_warranty'])
check('品牌汇总存库(哈轴)', sA['quote_brand'] == '哈轴', sA['quote_brand'])
check('行明细quote_details已存2行', bool(sA['quote_details']) and len(json.loads(sA['quote_details'])) == 2)
d = client.get('/api/inquiries/%d' % iid1).get_json()
sA_d = [s for s in d['suppliers'] if s['id'] == sidA][0]
check('详情接口返回交付日期', sA_d['quote_delivery'] == '7天', sA_d['quote_delivery'])
check('详情接口返回质保时间', sA_d['quote_warranty'] == '3个月', sA_d['quote_warranty'])

# 最后一家报价 → 自动定标审批(原V11.75 INSERT参数错误已修复)
r = client.post('/api/inquiry/vendor/%s/quote' % sups1[1]['token'], json={'quote_price': 300, 'quote_remark': '含税含运'})
check('商家乙报价(触发自动定标)', r.status_code == 200, r.get_json())
check('自动定标: 询价单状态=定标审批中', q1("SELECT status FROM inquiries WHERE id=?", (iid1,))[0] == '定标审批中', q1("SELECT status FROM inquiries WHERE id=?", (iid1,))[0])
check('自动定标: 草稿订单已生成', q1("SELECT status FROM purchase_orders WHERE inquiry_id=?", (iid1,))[0] == '草稿')
check('自动定标: 审批实例pending', q1("SELECT COUNT(*) FROM approval_instances WHERE biz_type='inquiry_approval' AND biz_id=? AND status='pending'", (iid1,))[0] == 1)
check('自动定标: 发起start_instances', CALLED['start_instances'] >= 1, CALLED['start_instances'])
check('自动定标: 询价审批记录=审批中', q1("SELECT status FROM inquiry_approvals WHERE inquiry_id=? ORDER BY id DESC LIMIT 1", (iid1,))[0] == '审批中')
check('无真实钉钉实例(已屏蔽)', len(q("SELECT 1 FROM dingtalk_instances WHERE biz_type='inquiry_approval'")) == 0)

# 审批中心列表
pend = client.get('/api/approvals/all-pending').get_json()
ip = [x for x in pend if x['biz_type'] == 'inquiry_approval' and x['biz_id'] == iid1]
check('审批中心显示询价定标待审+单号', len(ip) == 1 and ip[0]['biz_no'] == inq1, ip[0]['biz_no'] if ip else None)

# 领导审批通过
r = appmod.finish_approvals('inquiry_approval', iid1, 'ok', '测试领导', 1, '测试通过')
check('审批通过处理成功(无500/无锁)', r is True)
check('审批实例已approved', q1("SELECT COUNT(*) FROM approval_instances WHERE biz_type='inquiry_approval' AND biz_id=? AND status='approved'", (iid1,))[0] == 1)
check('询价单状态=已生成订单', q1("SELECT status FROM inquiries WHERE id=?", (iid1,))[0] == '已生成订单')
check('选中最低价家(甲230<乙300)', q1("SELECT selected_supplier_id FROM inquiries WHERE id=?", (iid1,))[0] == sidA)
check('询价审批记录=已完成', q1("SELECT status FROM inquiry_approvals WHERE inquiry_id=? ORDER BY id DESC LIMIT 1", (iid1,))[0] == '已完成')
orders = q("SELECT * FROM purchase_orders WHERE inquiry_id=?", (iid1,))
check('订单只有1张(草稿被完善不重复)', len(orders) == 1, len(orders))
check('订单已生效(已通过)', orders[0]['status'] == '已通过', orders[0]['status'])
check('订单明细2行按比例分摊', q1("SELECT COUNT(*) FROM order_items WHERE order_id=?", (orders[0]['id'],))[0] == 2)
sum_amt = q1("SELECT SUM(total_amount) FROM order_items WHERE order_id=?", (orders[0]['id'],))[0]
check('订单明细合计=报价230', abs(float(sum_amt) - 230.0) < 0.02, sum_amt)
oid1 = orders[0]['id']

# ══ 订单进度(问题6) ══
def order_progress(oid):
    for o in client.get('/api/orders').get_json():
        if o['id'] == oid:
            return o['progress']
    return None
check('订单进度: 未联系厂家(红)', order_progress(oid1) == 'contact', order_progress(oid1))
c = sqlite3.connect('data/purchase.db', timeout=30)
c.execute("UPDATE order_items SET status='已发货' WHERE order_id=?", (oid1,)); c.commit(); c.close()
check('订单进度: 已下单在途(黄)', order_progress(oid1) == 'shipped', order_progress(oid1))
c = sqlite3.connect('data/purchase.db', timeout=30)
c.execute("UPDATE order_items SET status='已到货' WHERE order_id=?", (oid1,)); c.commit(); c.close()
check('订单进度: 已到货(绿)', order_progress(oid1) == 'arrived', order_progress(oid1))
c = sqlite3.connect('data/purchase.db', timeout=30)
c.execute("UPDATE order_items SET status='未联系' WHERE order_id=?", (oid1,)); c.commit(); c.close()
check('订单进度: 草稿显示—', (lambda: (lambda o: o['progress'])([x for x in client.get('/api/orders').get_json() if x['id'] == oid1][0]))() == 'contact')

# ══ 询价2: SQL造两家报价 → 手动提交 → 驳回 → 恢复 → 重新提交(去重) ══
r = client.post('/api/inquiries', json={'req_id': rid2, 'suppliers': [
    {'name': 'V126厂家丙', 'contact': '丙', 'phone': '10000000003'},
    {'name': 'V126厂家丁', 'contact': '丁', 'phone': '10000000004'}]})
inq2 = r.get_json()['inq_no']
iid2 = q1("SELECT id FROM inquiries WHERE inq_no=?", (inq2,))['id']
sups2 = q("SELECT * FROM inquiry_suppliers WHERE inquiry_id=?", (iid2,))
c = sqlite3.connect('data/purchase.db', timeout=30)
c.execute("UPDATE inquiry_suppliers SET quote_price=111, quote_details=?, quote_delivery='10天', quote_warranty='1年', quote_time=datetime('now','localtime') WHERE id=?",
          (json.dumps([{'unit_price': 27.75, 'qty': 4, 'delivery': '10天', 'warranty': '1年', 'brand': '', 'remark': ''}]), sups2[0]['id']))
c.execute("UPDATE inquiry_suppliers SET quote_price=222, quote_time=datetime('now','localtime') WHERE id=?", (sups2[1]['id'],))
c.commit(); c.close()
before = CALLED['start_instances']
r = client.post('/api/inquiries/%d/submit' % iid2)
check('手动提交定标审批成功', r.status_code == 200 and r.get_json().get('success'), r.get_json())
check('手动提交发起start_instances(钉钉)', CALLED['start_instances'] == before + 1, CALLED['start_instances'])
check('手动提交: 询价单状态=定标审批中', q1("SELECT status FROM inquiries WHERE id=?", (iid2,))[0] == '定标审批中')
check('手动提交: 审批实例pending', q1("SELECT COUNT(*) FROM approval_instances WHERE biz_type='inquiry_approval' AND biz_id=? AND status='pending'", (iid2,))[0] == 1)

r = appmod.finish_approvals('inquiry_approval', iid2, 'reject', '测试领导', 1, '测试驳回')
check('驳回处理成功(无500/无锁)', r is True)
check('驳回后询价单恢复询价中', q1("SELECT status FROM inquiries WHERE id=?", (iid2,))[0] == '询价中', q1("SELECT status FROM inquiries WHERE id=?", (iid2,))[0])
check('驳回后审批记录=已驳回', q1("SELECT status FROM inquiry_approvals WHERE inquiry_id=? ORDER BY id DESC LIMIT 1", (iid2,))[0] == '已驳回')
check('驳回后审批实例=rejected', q1("SELECT COUNT(*) FROM approval_instances WHERE biz_type='inquiry_approval' AND biz_id=? AND status='rejected'", (iid2,))[0] == 1)
check('驳回后无新订单(草稿1张)', len(q("SELECT * FROM purchase_orders WHERE inquiry_id=?", (iid2,))) == 1)

r = client.post('/api/inquiries/%d/submit' % iid2)
check('驳回后重新提交成功', r.status_code == 200, r.get_json())
check('重新提交不重复建草稿', len(q("SELECT * FROM purchase_orders WHERE inquiry_id=?", (iid2,))) == 1)

# ══ 轮询保存钉钉表单值(选定供应商) ══
c = sqlite3.connect('data/purchase.db', timeout=30)
c.execute("INSERT INTO dingtalk_instances(instance_code,biz_type,biz_id,status) VALUES('V126FAKE1','inquiry_approval',?, 'pending')", (iid1,))
c.commit(); c.close()
fake_inst = {'status': 'APPROVED', 'result': 'agree',
             'form_component_values': [{'name': '选定供应商', 'value': json.dumps([sidB])}]}
orig_query = appmod.dt_query_instance
appmod.dt_query_instance = lambda iid: fake_inst
appmod.dt_poll_results()
appmod.dt_query_instance = orig_query
fv = q1("SELECT form_values FROM dingtalk_instances WHERE instance_code='V126FAKE1'")
check('轮询保存钉钉表单值(选定供应商)', fv is not None and '选定供应商' in (fv[0] or ''), (fv[0] if fv else '')[:80])
c = sqlite3.connect('data/purchase.db', timeout=30)
c.execute("DELETE FROM dingtalk_instances WHERE instance_code='V126FAKE1'"); c.commit(); c.close()

# ══ 导出Excel: 交付/质保列 ══
from openpyxl import load_workbook
import io
resp = client.get('/api/inquiries/%d/export' % iid1)
check('导出Excel 200', resp.status_code == 200, resp.status_code)
ws = load_workbook(io.BytesIO(resp.data)).active
def find_row_containing(text):
    for r in range(1, 40):
        for cc in range(1, 25):
            if str(ws.cell(row=r, column=cc).value or '') == text:
                return r
    return None
hdr_row = find_row_containing('V126厂家甲 单价')
check('找到表头行', hdr_row is not None, hdr_row)
headers = [ws.cell(row=hdr_row, column=ci).value for ci in range(1, 17)]
hdr_txt = '|'.join(str(x or '') for x in headers)
check('Excel表头含 交付 列', 'V126厂家甲 交付' in hdr_txt, hdr_txt)
check('Excel表头含 质保 列', 'V126厂家甲 质保' in hdr_txt, hdr_txt)
check('Excel表头含 备注 列', 'V126厂家甲 备注' in hdr_txt, hdr_txt)
data_row = hdr_row + 1
row_vals = [ws.cell(row=data_row, column=ci).value for ci in range(1, 17)]
check('Excel数据行: 甲单价30', row_vals[4] == 30, row_vals[4])
check('Excel数据行: 甲交付7天', row_vals[7] == '7天', row_vals[7])
check('Excel数据行: 甲质保3个月', row_vals[8] == '3个月', row_vals[8])
check('Excel数据行: 乙(旧版总价)交付为空', row_vals[13] in (None, ''), row_vals[13])
total_row = find_row_containing('合计')
check('找到合计行', total_row is not None, total_row)
check('Excel合计行甲总价230', float(ws.cell(row=total_row, column=6).value or 0) == 230.0, ws.cell(row=total_row, column=6).value)

# ══ 清理 ══
c = sqlite3.connect('data/purchase.db', timeout=30)
rids = [r0['id'] for r0 in q("SELECT id FROM purchase_requests WHERE purpose LIKE ?", ('%' + MARK + '%',))]
iids = [r0['id'] for r0 in q("SELECT id FROM inquiries WHERE req_id IN (%s)" % ','.join('?' * len(rids)), tuple(rids))] if rids else []
for iid in iids:
    c.execute("DELETE FROM inquiry_suppliers WHERE inquiry_id=?", (iid,))
    c.execute("DELETE FROM inquiry_approvals WHERE inquiry_id=?", (iid,))
    c.execute("DELETE FROM approval_instances WHERE biz_type='inquiry_approval' AND biz_id=?", (iid,))
    c.execute("DELETE FROM dingtalk_instances WHERE biz_type='inquiry_approval' AND biz_id=?", (iid,))
    c.execute("DELETE FROM inquiries WHERE id=?", (iid,))
    oids = [r0['id'] for r0 in q("SELECT id FROM purchase_orders WHERE inquiry_id=?", (iid,))]
    for oid in oids:
        c.execute("DELETE FROM order_items WHERE order_id=?", (oid,))
    if oids:
        c.execute("DELETE FROM purchase_orders WHERE id IN (%s)" % ','.join('?' * len(oids)), tuple(oids))
for rid in rids:
    c.execute("DELETE FROM request_items WHERE req_id=?", (rid,))
    c.execute("DELETE FROM purchase_requests WHERE id=?", (rid,))
for tbl in ('inquiries', 'inquiry_suppliers', 'inquiry_approvals', 'purchase_orders', 'order_items', 'approval_instances', 'dingtalk_instances', 'purchase_requests', 'request_items'):
    c.execute("UPDATE sqlite_sequence SET seq=COALESCE((SELECT MAX(id) FROM %s),0) WHERE name=?" % tbl, (tbl,))
c.commit(); c.close()
leftover = q("SELECT COUNT(*) FROM purchase_requests WHERE purpose LIKE ?", ('%' + MARK + '%',))[0][0]
check('测试数据已清理干净', leftover == 0, leftover)

print()
print('PASS: %d  FAIL: %d' % (len(PASS), len(FAIL)))
sys.exit(1 if FAIL else 0)
