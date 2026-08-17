# -*- coding: utf-8 -*-
"""把桌面 .xls 模板精确转换为 .xlsx 系统模板（保留字体/边框/对齐/列宽/行高/合并单元格）"""
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

H_ALIGN = {0: 'general', 1: 'left', 2: 'center', 3: 'right', 4: 'fill', 5: 'justify', 6: 'centerContinuous', 7: 'distributed'}
V_ALIGN = {0: 'top', 1: 'center', 2: 'bottom', 3: 'justify', 4: 'distributed'}

def convert(src_xls, dst_xlsx):
    book = xlrd.open_workbook(src_xls, formatting_info=True)
    sh = book.sheet_by_index(0)
    wb = Workbook(); ws = wb.active
    ws.title = sh.name if sh.name else 'Sheet1'
    thin = Side(style='thin', color='000000')

    # 列宽 (xlrd width 单位 1/256 字符)
    for c, info in sh.colinfo_map.items():
        if info.width:
            ws.column_dimensions[get_column_letter(c + 1)].width = info.width / 256.0
    # 行高 (xlrd 单位 twips, 1pt=20twips)
    for r, info in sh.rowinfo_map.items():
        if info.height:
            ws.row_dimensions[r + 1].height = info.height / 20.0

    for r in range(sh.nrows):
        for c in range(sh.ncols):
            v = sh.cell_value(r, c)
            if v == '' or v is None:
                continue
            cell = ws.cell(row=r + 1, column=c + 1)
            cell.value = v
            # 格式
            try:
                xf_idx = sh.cell_xf_index(r, c)
                xf = book.xf_list[xf_idx]
                # 字体
                f = book.font_list[xf.font_index]
                cell.font = Font(name=f.name, size=f.height / 20.0, bold=bool(f.bold), italic=bool(f.italic))
                # 边框
                def _side(ls):
                    if ls in (1, 2, 3, 4, 7): return Side(style='thin', color='000000')
                    if ls in (5,): return Side(style='medium', color='000000')
                    return None
                cell.border = Border(
                    left=_side(xf.border_left_line_style), right=_side(xf.border_right_line_style),
                    top=_side(xf.border_top_line_style), bottom=_side(xf.border_bottom_line_style))
                # 对齐
                al = xf.alignment
                cell.alignment = Alignment(horizontal=H_ALIGN.get(al.hor_align, 'general'),
                                           vertical=V_ALIGN.get(al.vert_align, 'bottom'),
                                           wrap_text=bool(al.text_wrapped))
                # 背景色
                if xf.background.pattern_colour_index and xf.background.pattern_colour_index not in (64, 65):
                    try:
                        rgb = book.colour_map.get(xf.background.pattern_colour_index)
                        if rgb:
                            cell.fill = PatternFill('solid', fgColor='%02X%02X%02X' % rgb[:3])
                    except Exception:
                        pass
            except Exception:
                pass

    # 合并单元格 (rlo, rhi, clo, chi)
    for rlo, rhi, clo, chi in sh.merged_cells:
        if rhi > rlo + 1 or chi > clo + 1:
            ws.merge_cells(start_row=rlo + 1, start_column=clo + 1, end_row=rhi, end_column=chi)

    wb.save(dst_xlsx)
    print(f'转换完成: {src_xls} → {dst_xlsx} ({sh.nrows}行x{sh.ncols}列)')


if __name__ == '__main__':
    base = r'C:\Users\mjj\Desktop'
    dst = r'C:\Users\mjj\Desktop\purchase_system\data'
    convert(base + r'\入库单(3).xls', dst + r'\入库单模板.xlsx')
    convert(base + r'\出库单(1)(1).xls', dst + r'\出库单模板.xlsx')
