# -*- coding: utf-8 -*-
"""V11.4 采购订单下载Excel验证"""
import os, sys, sqlite3, json, io
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import app as S
S.start_instances = lambda *a, **k: None

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'purchase.db')
conn = sqlite3.connect(DB, timeout=30)
MARK = '【订单下载验证】'

# 清理
oids = [r[0] for r in conn.execute("SELECT id FROM purchase_orders WHERE remark LIKE ?", ('%' + MARK + '%',)).fetchall()]
for (oid,) in [(o,) for o in oids]:
    conn.execute("DELETE FROM order_items WHERE order_id=?", (oid,))
    conn.execute("DELETE FROM approval_instances WHERE biz_type='purchase_order' AND biz_id=?", (oid,))
conn.execute("DELETE FROM purchase_orders WHERE remark LIKE ?", ('%' + MARK + '%',))
conn.commit()

client = S.app.test_client()
r = client.post('/api/login', json={'username': 'admin', 'password': 'admin123'})
assert r.status_code == 200, '登录失败'

# 创建订单(带自定义交易模式+明细2项)
r = client.post('/api/orders', json={
    'items': [
        {'item_name': '钢材', 'spec': 'Q235 10mm', 'unit': '吨', 'quantity': 2, 'price': 4500, 'tax_rate': 13},
        {'item_name': '螺栓', 'spec': 'M12', 'unit': '个', 'quantity': 100, 'price': 2, 'tax_rate': 13},
    ],
    'supplier': '甲供应商', 'requester': '综合办', 'category': '备品备件',
    'trade_mode': '月结30天', 'remark': MARK + ' 询价选中: 甲供应商 报价¥10100', 'target_date': '2026-09-01'})
assert r.status_code == 200, r.get_data(as_text=True)
oid = r.get_json()['id']
order_no = r.get_json()['order_no']
print('① 创建订单 OK:', order_no, 'id=', oid)

# 下载
r = client.get('/api/orders/%d/download' % oid)
assert r.status_code == 200, r.get_data(as_text=True)[:300]
data = r.data
assert data[:2] == b'PK', '不是xlsx文件'
print('② 下载Excel OK, 大小:', len(data))

# 内容检查
from openpyxl import load_workbook
wb = load_workbook(io.BytesIO(data))
ws = wb['采购订单']
rows = [[c.value for c in row] for row in ws.iter_rows()]
flat = [str(v) for v in [x for row in rows for x in row] if v is not None]
joined = '|'.join(flat)
checks = {
    '标题': '采 购 订 单' in joined,
    '订单编号': order_no in joined,
    '供应商': '甲供应商' in joined,
    '交易模式': '月结30天' in joined,
    '明细1钢材': '钢材' in joined,
    '明细2螺栓': '螺栓' in joined,
    '备注': MARK in joined,
    '合计': '合计' in joined,
    '审批区': '审批进度' in joined or '签字' in joined,
}
print('③ Excel内容检查:')
for k, v in checks.items():
    print('   ', '✅' if v else '❌', k)

# 金额核对: 找"合计"行, 金额列=10396.00 (2*4500*1.13 + 100*2*1.13 = 10170+226 = 10396)
amt_ok = False
for row in rows:
    if row and row[3] == '合计':
        amt_ok = abs(float(row[7] or 0) - 10396.0) < 1
        break
print('    ', '✅' if amt_ok else '❌', '金额合计 10396.00')

all_pass = all(checks.values()) and amt_ok
print('=== 结果:', 'ALL PASS' if all_pass else 'FAIL', '===')

# 清理
conn.execute("DELETE FROM order_items WHERE order_id=?", (oid,))
conn.execute("DELETE FROM approval_instances WHERE biz_type='purchase_order' AND biz_id=?", (oid,))
conn.execute("DELETE FROM purchase_orders WHERE id=?", (oid,))
conn.execute("DELETE FROM logs WHERE detail LIKE ?", ('%订单下载验证%',))
conn.commit(); conn.close()
print('清理完成')
sys.exit(0 if all_pass else 1)
